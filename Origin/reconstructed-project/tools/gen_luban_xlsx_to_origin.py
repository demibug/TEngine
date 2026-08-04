#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""将 unity-export/config 与 origin_project/data 的 JSON 配置转为 Luban xlsx 数据表。

本脚本输出到 Origin/config/，使该目录成为一份完整的、可供 Unity 工程 Luban 直接使用的配置源目录：
  - 生成全部业务数据表 xlsx（含 ai-difficulty 单行表）
  - 拷贝 schema 文件（__tables__/__beans__/__enums__、item.xlsx、luban.conf）并补注册 ai-difficulty 行
  - 保留已存在的 rankData.xlsx / rankDataPlayer.xlsx（本脚本不生成、不覆盖）

重跑幂等：数据表覆盖重写，__tables__ 注册行按 full_name 去重追加。
注意：rank.xlsx / weapon.xlsx / weaponTxt.xlsx 用重构版覆盖（依据用户决策）。
"""
import json
import os
import shutil
from openpyxl import Workbook, load_workbook

# 源数据目录
EXPORT_CFG = r'E:\MyWork\MyTD\TEngine\Origin\reconstructed-project\unity-export\config'
ORIGIN_DATA = r'E:\MyWork\MyTD\TEngine\Origin\reconstructed-project\origin_project\data'
# Unity Luban 源目录（schema 文件来源）
UNITY_DATAS = r'E:\MyWork\MyTD\TEngine\Configs\GameConfig\Datas'
UNITY_CONF = r'E:\MyWork\MyTD\TEngine\Configs\GameConfig\luban.conf'
# 目标目录
OUT_DIR = r'E:\MyWork\MyTD\TEngine\Origin\config'


def load_json(d, name):
    with open(os.path.join(d, name), 'r', encoding='utf-8') as f:
        return json.load(f)


def num(v):
    """数值格式化：float 去浮点尾，int/bool 直转，None 返回空串。"""
    if v is None:
        return ''
    if isinstance(v, bool):
        return 'true' if v else 'false'
    if isinstance(v, float):
        return f"{v:g}"
    return str(v)


def fmt_list(vals, sep=','):
    """一维 list 转单元格字符串。"""
    if not vals:
        return ''
    return sep.join(num(v) for v in vals)


def fmt_vec2_list(points, sep=';'):
    """坐标点列表转单元格字符串，复用内置 vector2int(x,y)。"""
    if not points:
        return ''
    return sep.join(f"{num(p['x'])},{num(p['y'])}" for p in points)


def fmt_2d_list(matrix, outer=';', inner=','):
    """二维 list 转单元格字符串。"""
    if not matrix:
        return ''
    return outer.join(inner.join(num(v) for v in row) for row in matrix)


def opt(d, key):
    """可空字段：缺失或 None 返回空串。"""
    return num(d.get(key))


def fmt_timeline(tl):
    """boss.Timeline bean 转单元格字符串，sep=,。"""
    if not tl:
        return ''
    return f"{num(tl.get('effectAtMs'))},{num(tl.get('completeAtMs'))}"


def data_rows(records):
    """多行表数据：每行前补空串作行标识列。"""
    return [[''] + r for r in records]


# ---- 加载全部源 JSON ----
units = load_json(EXPORT_CFG, 'units.json')
weapons_reg = load_json(EXPORT_CFG, 'weapons.json')
bosses = load_json(EXPORT_CFG, 'bosses.json')
buffs = load_json(EXPORT_CFG, 'buffs.json')
skills = load_json(EXPORT_CFG, 'skills.json')
enemies = load_json(EXPORT_CFG, 'enemies.json')
generals = load_json(EXPORT_CFG, 'generals.json')
waves = load_json(EXPORT_CFG, 'waves.json')
maps = load_json(EXPORT_CFG, 'maps.json')
economy = load_json(EXPORT_CFG, 'battle-economy.json')
projectiles = load_json(EXPORT_CFG, 'projectiles.json')
events = load_json(EXPORT_CFG, 'events.json')
result_schema = load_json(EXPORT_CFG, 'battle-result-schema.json')
rank = load_json(ORIGIN_DATA, 'rank.json')
weapon = load_json(ORIGIN_DATA, 'weapon.json')
weapon_txt = load_json(ORIGIN_DATA, 'weaponTxt.json')
ai_diff = load_json(EXPORT_CFG, 'ai-difficulty.json')

# ---- 表定义 ----
# 每项: file, full_name, value_type, mode(''=map/'one'), comment, fields[(name,type,comment)], rows
TABLES = []


def add(file, full_name, vt, mode, comment, fields, rows):
    TABLES.append(dict(file=file, full_name=full_name, value_type=vt,
                       mode=mode, comment=comment, fields=fields, rows=rows))


# 1. 小兵基础配置(多行)
add('unit.xlsx', 'battle.TbUnit', 'Unit', '', '小兵基础配置',
    [('index', 'int', '索引'), ('text', 'string', '显示名'), ('animationKey', 'string', '动画键'),
     ('rangeCells', 'float', '攻击距离(格)'), ('attackDamage', 'int', '攻击力'),
     ('attackIntervalSeconds', 'float', '攻击间隔(秒)'), ('damageMode', 'string', '伤害模式'),
     ('targetPolicy', 'string', '目标策略')],
    data_rows([[u['index'], u['text'], u['animationKey'], num(u['rangeCells']),
                u['attackDamage'], num(u['attackIntervalSeconds']), u['damageMode'], u['targetPolicy']]
               for u in units['units']]))

# 2. 武器符号注册表(多行,逆向快照)
add('weapon_registry.xlsx', 'battle.TbWeaponRegistry', 'WeaponRegistry', '', '武器符号注册表(逆向快照)',
    [('symbol', 'string', '符号(主键)'), ('name', 'string', '名称'), ('type', 'int', '类型'),
     ('index', 'string', '索引'), ('status', 'string?', '状态'), ('buffDependency', 'string?', 'Buff依赖')],
    data_rows([[w['symbol'], w['name'], w['type'], w['index'], opt(w, 'status'), opt(w, 'buffDependency')]
               for w in weapons_reg['weapons']]))

# 3. Boss配置(多行)
add('boss.xlsx', 'battle.TbBoss', 'Boss', '', 'Boss配置',
    [('key', 'string', '键(主键)'), ('name', 'string', '名称'), ('originalSymbol', 'string', '原始符号'),
     ('sourceRange', 'string', '源码区间'), ('skillKey', 'string', '技能键'), ('animationKey', 'string', '动画键'),
     ('resourcePath', 'string', '资源路径'), ('attackAnimation', 'string', '攻击动画'),
     ('followupAnimation', 'string?', '后续动画'), ('idleAnimation', 'string', '待机动画'),
     ('timeline', 'boss.Timeline', '时间轴')],
    data_rows([[b['key'], b['name'], b['originalSymbol'], b['sourceRange'], b['skillKey'],
                b['animationKey'], b['resourcePath'], b['attackAnimation'], opt(b, 'followupAnimation'),
                b['idleAnimation'], fmt_timeline(b.get('timeline'))]
               for b in bosses['bosses']]))

# 4. Buff类型配置(多行)
add('buff.xlsx', 'battle.TbBuff', 'Buff', '', 'Buff类型配置',
    [('type', 'int', '类型(主键)'), ('name', 'string', '名称'), ('label', 'string?', '中文标签'),
     ('kind', 'int', '类别'), ('channels', '(list#sep=,),int', '通道列表')],
    data_rows([[b['type'], b['name'], opt(b, 'label'), b['kind'], fmt_list(b.get('channels'))]
               for b in buffs['buffs']]))

# 5. 技能配置(多行)
add('skill.xlsx', 'battle.TbSkill', 'Skill', '', '技能配置',
    [('key', 'string', '键(主键)'), ('name', 'string', '名称'), ('category', 'string', '类别'),
     ('description', 'string', '描述'), ('healthMultiplier', 'int?', '血量倍数'),
     ('speed', 'int?', '速度'), ('rangeTiles', 'float?', '范围(格)'),
     ('cooldownSeconds', 'int?', '冷却(秒)'), ('source', 'string', '源码标记'), ('confidence', 'string?', '置信度')],
    data_rows([[s['key'], s['name'], s['category'], s['description'], opt(s, 'healthMultiplier'),
                opt(s, 'speed'), opt(s, 'rangeTiles'), opt(s, 'cooldownSeconds'), s['source'], opt(s, 'confidence')]
               for s in skills['skills']]))

# 6. 敌人类型配置(多行,逆向快照)
add('enemy.xlsx', 'battle.TbEnemy', 'Enemy', '', '敌人类型配置(逆向快照)',
    [('key', 'string', '键(主键)'), ('symbol', 'string', '符号'), ('status', 'string?', '状态'),
     ('resource', 'string?', '资源路径'), ('deferred', 'string?', '延迟项'),
     ('levelMultipliers', '(list#sep=,),float?', '等级倍数')],
    data_rows([[e['key'], e['symbol'], opt(e, 'status'), opt(e, 'resource'), opt(e, 'deferred'),
                fmt_list(e.get('levelMultipliers'))]
               for e in enemies['types']]))

# 7. 武将配置(多行)
add('general.xlsx', 'battle.TbGeneral', 'General', '', '武将配置',
    [('index', 'int', '索引(主键)'), ('name', 'string', '名称'), ('family', 'string', '姓氏'),
     ('partWords', '(list#sep=,),string', '名字拆字'), ('weaponType', 'int', '武器类型'), ('status', 'string?', '状态')],
    data_rows([[g['index'], g['name'], g['family'], fmt_list(g.get('partWords')), g['weaponType'], opt(g, 'status')]
               for g in generals['generals']]))

# 8. 军衔关卡配置(多行) —— 用重构版覆盖 Origin/config/rank.xlsx
add('rank.xlsx', 'battle.TbRank', 'Rank', '', '军衔关卡配置',
    [('id', 'int', 'ID(主键)'), ('rank', 'string', '军衔'), ('level', 'int', '等级'),
     ('reward', 'int', '奖励'), ('addHp', 'int', '加血'),
     ('weapons', '(list#sep=,),int', '武器配置(0-4)'), ('difficulties', '(list#sep=,),int', '难度(0-3)'),
     ('map', 'string', '地图')],
    data_rows([[r['id'], r['rank'], r['level'], r['reward'], r['addHp'],
                fmt_list([r.get('weapon0'), r.get('weapon1'), r.get('weapon2'), r.get('weapon3'), r.get('weapon4')]),
                fmt_list([r.get('difficulty0'), r.get('difficulty1'), r.get('difficulty2'), r.get('difficulty3')]),
                r['map']]
               for r in rank]))

# 9. 武器配置(多行,原始) —— 用重构版覆盖 Origin/config/weapon.xlsx
add('weapon.xlsx', 'battle.TbWeapon', 'Weapon', '', '武器配置(原始)',
    [('id', 'int', 'ID(主键)'), ('type', 'int', '类型'), ('txt', 'string', '名称'),
     ('rarity', 'int', '稀有度'), ('rareTxt', 'string?', '稀有字'), ('addAttPower', 'int', '附加攻击力'),
     ('exclusive', 'string?', '专属'), ('scale', 'float', '缩放'), ('anchorY', 'float', '锚点Y'),
     ('intro', 'string', '简介'), ('fragmentNum', 'int', '碎片数')],
    data_rows([[w['id'], w['type'], w['txt'], w['rarity'], opt(w, 'rareTxt'), w['addAttPower'],
                opt(w, 'exclusive'), num(w['scale']), num(w['anchorY']), w['intro'], w['fragmentNum']]
               for w in weapon]))

# 10. 武器名字拆字(多行) —— 用重构版覆盖 Origin/config/weaponTxt.xlsx（文件名对齐原始）
add('weaponTxt.xlsx', 'battle.TbWeaponText', 'WeaponText', '', '武器名字拆字',
    [('id', 'int', 'ID(主键)'), ('txt', 'string', '单字'), ('quality', 'int', '品质')],
    data_rows([[w['id'], w['txt'], w['quality']] for w in weapon_txt]))

# 11. 事件名映射(多行)
add('event.xlsx', 'battle.TbEvent', 'Event', '', '事件名映射',
    [('eventName', 'string', '事件名(主键)'), ('code', 'string', '事件码')],
    data_rows([[k, v] for k, v in events.items()]))

# 12. 战斗结果字段说明(多行)
add('result_schema.xlsx', 'battle.TbResultSchema', 'ResultSchema', '', '战斗结果字段说明',
    [('field', 'string', '字段(主键)'), ('typeDesc', 'string', '类型描述')],
    data_rows([[k, v] for k, v in result_schema.items()]))

# 13. 小兵等级倍数(单行)
add('unit_level.xlsx', 'battle.TbUnitLevel', 'UnitLevel', 'one', '小兵等级倍数',
    [('maxLevel', 'int', '最大等级'),
     ('damageLevelMultipliers', '(list#sep=,),float', '伤害等级倍数'),
     ('attackSpeedLevelMultipliers', '(list#sep=,),float', '攻速等级倍数')],
    [['', units['maxLevel'], fmt_list(units['damageLevelMultipliers']), fmt_list(units['attackSpeedLevelMultipliers'])]])

# 14. 波次配置(单行)
add('wave.xlsx', 'battle.TbWave', 'Wave', 'one', '波次配置',
    [('waveUnitCounts', '(list#sep=,),int', '每波怪物数'),
     ('bossWaveNumbers', '(list#sep=,),int', 'Boss波次号'),
     ('bossSpawnChances', '(list#sep=,),float', 'Boss出现概率'),
     ('spawnStrategyWeights', '(list#sep=,),int', '生成策略权重'),
     ('spawnStrategies', '(list#sep=;),(list#sep=,),float', '生成策略表')],
    [['', fmt_list(waves['waveUnitCounts']), fmt_list(waves['bossWaveNumbers']),
      fmt_list(waves['bossSpawnChances']), fmt_list(waves['spawnStrategyWeights']),
      fmt_2d_list(waves['spawnStrategies'])]])

# 15. 地图配置(单行)
add('map.xlsx', 'battle.TbMap', 'Map', 'one', '地图配置',
    [('gridWidth', 'int', '网格宽'), ('gridHeight', 'int', '网格高'),
     ('width', 'int', '宽'), ('height', 'int', '高'), ('mapIndex', 'int', '地图索引'),
     ('blocks', 'string?', '阻挡点'), ('playerPath', '(list#sep=;),vector2int', '玩家路径'),
     ('opponentPath', '(list#sep=;),vector2int', '对手路径')],
    [['', maps['gridWidth'], maps['gridHeight'], maps['width'], maps['height'], maps['mapIndex'],
      num(maps.get('blocks')), fmt_vec2_list(maps['playerPath']), fmt_vec2_list(maps['opponentPath'])]])

# 16. 战斗经济配置(单行)
add('economy.xlsx', 'battle.TbEconomy', 'Economy', 'one', '战斗经济配置',
    [('initialGold', 'int', '初始金币'), ('refreshCostStart', 'int', '刷新起始消耗'),
     ('refreshCostIncrement', 'int', '刷新递增'), ('unitBaseCost', 'int', '单位基础消耗'),
     ('handSize', 'int', '手牌数')],
    [['', economy['initialGold'], economy['refreshCostStart'], economy['refreshCostIncrement'],
      economy['unitBaseCost'], economy['handSize']]])

# 17. 弹道类型(单行)
add('projectile.xlsx', 'battle.TbProjectile', 'Projectile', 'one', '弹道类型',
    [('types', '(list#sep=,),string', '弹道类型列表')],
    [['', fmt_list(projectiles['types'])]])

# 18. AI难度配置(单行,ai-advanced-strategy 变新增)
add('ai_difficulty.xlsx', 'battle.TbAiDifficulty', 'AiDifficulty', 'one', 'AI难度配置(4级动态难度)',
    [('decisionIntervalMs', '(list#sep=,),int', '决策间隔ms(0-3档)'),
     ('ni', '(list#sep=,),float', '快速结束概率(0-3档)'),
     ('ri', '(list#sep=,),float', 'XG触发概率(0-3档)'),
     ('hi', 'int', '刷牌阈值'),
     ('ii', '(list#sep=;),(list#sep=,),int', '周期收入表[4档][6波]'),
     ('ei', '(list#sep=,),int', '波次表'),
     ('oi', '(list#sep=,),int', '金币返还(0-3档)'),
     ('itemCooldownMs', 'int', '道具冷却ms')],
    [['', fmt_list(ai_diff['decisionIntervalMs']), fmt_list(ai_diff['ni']), fmt_list(ai_diff['ri']),
      ai_diff['hi'], fmt_2d_list(ai_diff['ii']), fmt_list(ai_diff['ei']), fmt_list(ai_diff['oi']),
      ai_diff['itemCooldownMs']]])


# ---- 写数据表 xlsx ----
def write_table(t):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(['##var'] + [f[0] for f in t['fields']])
    ws.append(['##type'] + [f[1] for f in t['fields']])
    ws.append(['##group'] + ['' for _ in t['fields']])
    ws.append(['##'] + [f[2] for f in t['fields']])
    for row in t['rows']:
        ws.append(row)
    wb.save(os.path.join(OUT_DIR, t['file']))


# ---- 拷贝 schema 与 item.xlsx、luban.conf ----
def copy_schema():
    """拷贝 __tables__/__beans__/__enums__/item.xlsx/luban.conf 到 OUT_DIR。"""
    for fn in ['__tables__.xlsx', '__beans__.xlsx', '__enums__.xlsx', 'item.xlsx']:
        shutil.copy2(os.path.join(UNITY_DATAS, fn), os.path.join(OUT_DIR, fn))
    shutil.copy2(UNITY_CONF, os.path.join(OUT_DIR, 'luban.conf'))


# ---- 追加 __tables__ 注册行(去重) ----
def register_tables():
    TBL_PATH = os.path.join(OUT_DIR, '__tables__.xlsx')
    wb = load_workbook(TBL_PATH)
    ws = wb.active
    existing = set()
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[1]:
            existing.add(row[1])
    r = ws.max_row + 1
    new_count = 0
    for t in TABLES:
        if t['full_name'] in existing:
            continue
        ws.cell(r, 1, '')
        ws.cell(r, 2, t['full_name'])
        ws.cell(r, 3, t['value_type'])
        ws.cell(r, 4, 'True')
        ws.cell(r, 5, t['file'])
        ws.cell(r, 6, '')
        ws.cell(r, 7, t['mode'])
        ws.cell(r, 8, '')
        ws.cell(r, 9, t['comment'])
        r += 1
        new_count += 1
    wb.save(TBL_PATH)
    return new_count


# ---- 追加 __beans__ 注册 boss.Timeline(去重) ----
def register_timeline_bean():
    BEAN_PATH = os.path.join(OUT_DIR, '__beans__.xlsx')
    wb = load_workbook(BEAN_PATH)
    ws = wb.active
    has_timeline = False
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[1] == 'boss.Timeline':
            has_timeline = True
            break
    if not has_timeline:
        r = ws.max_row + 1
        # 列: A##var B full_name C parent D valueType E sep F alias G comment H group I tags
        #     J name K alias L type M group N comment O tags P variants
        ws.cell(r, 1, '')
        ws.cell(r, 2, 'boss.Timeline')
        ws.cell(r, 5, ',')
        ws.cell(r, 7, 'Boss技能时间轴')
        ws.cell(r, 10, 'effectAtMs')
        ws.cell(r, 12, 'int')
        ws.cell(r, 14, '特效触发毫秒')
        r += 1
        ws.cell(r, 1, '')
        ws.cell(r, 10, 'completeAtMs')
        ws.cell(r, 12, 'int')
        ws.cell(r, 14, '完成毫秒')
        wb.save(BEAN_PATH)
        return True
    return False


# ---- 主流程 ----
if __name__ == '__main__':
    os.makedirs(OUT_DIR, exist_ok=True)
    copy_schema()
    for t in TABLES:
        write_table(t)
    new_tbl = register_tables()
    new_bean = register_timeline_bean()

    print('生成完成：')
    print(f'  输出目录: {OUT_DIR}')
    print(f'  数据表: {len(TABLES)} 张')
    for t in TABLES:
        print(f'    {t["file"]:24s} -> {t["full_name"]} ({len(t["rows"])}行) mode={t["mode"] or "map"}')
    print(f'  __tables__ 新增注册行: {new_tbl}')
    print(f'  __beans__ 新增 boss.Timeline: {"是" if new_bean else "已存在跳过"}')
    print('  保留文件: rankData.xlsx / rankDataPlayer.xlsx (本脚本不动)')
    print('  schema 拷贝: __tables__/__beans__/__enums__/item.xlsx/luban.conf')
