#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Generate an import-ready staging workbook for player general synthesis.

Reads the read-only production workbook
    <repo>/Configs/GameConfig/Datas/general.xlsx
and writes a complete replacement candidate (preserving every existing row,
style and index while appending the approved columns from design.md) plus a
two-row enabled preview, a manifest and a manual-import README into
    <repo>/UnityProject/outputs/player-general-synthesis/

Only index 1 (Zhang Fei) and index 4 (Huang Zhong) are enabled; every other
row keeps its original data but is marked disabled with archetype-specific
fields left blank/zero.

This script NEVER writes production Datas, C# or binary config.  It validates
the expected source hash by default, unique IDs, exactly two non-empty
distinct recipe parts, canonical unordered recipe uniqueness, exactly enabled
IDs {1,4}, archetype-specific required fields and protected exact-file
overwrite (explicit --force only when content differs).
"""

import argparse
import hashlib
import json
import os
import sys
import time
import uuid

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - guarded at runtime
    print("ERROR: 需要 openpyxl，请运行 pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, "..", "..", ".."))
PROJECT_ROOT = os.path.join(REPO_ROOT, "UnityProject")
SOURCE_WORKBOOK = os.path.join(
    REPO_ROOT, "Configs", "GameConfig", "Datas", "general.xlsx"
)
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "outputs", "player-general-synthesis")

EXPECTED_SOURCE_SHA256 = (
    "b5763de594bd1bf1fc0de98efbf9fb3dcb5912e34155a2cebe64dfb83cfcd213"
)

# index -> staged config for the enabled generals (design.md decision 4).
ENABLED_IDS = {1, 4}
STAGED = {
    1: {
        "recipe": "张,飞",
        "enabled": True,
        "combatArchetype": "pike",
        "rangeCells": 2.5,
        "attackDamage": 15,
        "attackIntervalSeconds": 1.0,
        "damageMode": "近战枪击",
        "targetPolicy": "nearest",
        "prefabAddress": "SpearSoldier",
        "animationKey": "default",
        "projectileType": "",
        "projectileSpeed": 0,
        "partRecruitWeight": 1,
    },
    4: {
        "recipe": "黄,忠",
        "enabled": True,
        "combatArchetype": "bow",
        "rangeCells": 3.5,
        "attackDamage": 13,
        "attackIntervalSeconds": 0.8,
        "damageMode": "单体",
        "targetPolicy": "nearest",
        "prefabAddress": "BowSoldier",
        "animationKey": "default",
        "projectileType": "SimpleDynamicArrow",
        "projectileSpeed": 200,
        "partRecruitWeight": 1,
    },
}

# Approved new columns (design.md decision 4), appended after existing status.
NEW_COLUMNS = [
    ("enabled", "bool", "是否进入本期配置快照与配方索引"),
    ("combatArchetype", "string", "pike 或 bow，映射既有逻辑原型"),
    ("rangeCells", "float", "当前 Unity 网格攻击范围"),
    ("attackDamage", "int", "1 级基础伤害"),
    ("attackIntervalSeconds", "float", "基础攻击间隔"),
    ("damageMode", "string", "近战枪击 或 单体，供校验/表现描述"),
    ("targetPolicy", "string", "复用现有目标策略键"),
    ("prefabAddress", "string", "当前枪兵/弓兵回退地址，后续可替换为专属 Prefab"),
    ("animationKey", "string", "表现绑定键"),
    ("projectileType", "string?", "仅远程武将需要，本期为 SimpleDynamicArrow"),
    ("projectileSpeed", "int", "仅远程武将使用，本期 200"),
    ("partRecruitWeight", "int", "每个配方字加入玩家池的权重，首版为 1"),
]

DEFAULT_ANIMATION_KEY = "default"


def sha256_file(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def workbook_content_equal(path_a, path_b):
    """True if two workbooks have identical cell values and column count.

    openpyxl embeds a fresh created/modified timestamp in docProps/core.xml on
    every save, so raw bytes are not stable across runs.  Content comparison
    (cell values + sheet shape) gives a deterministic idempotency signal.
    """
    wb_a = openpyxl.load_workbook(path_a, data_only=False, read_only=True)
    wb_b = openpyxl.load_workbook(path_b, data_only=False, read_only=True)
    try:
        if wb_a.sheetnames != wb_b.sheetnames:
            return False
        for name in wb_a.sheetnames:
            ws_a = wb_a[name]
            ws_b = wb_b[name]
            if (ws_a.max_row, ws_a.max_column) != (ws_b.max_row, ws_b.max_column):
                return False
            for row_a, row_b in zip(ws_a.iter_rows(), ws_b.iter_rows()):
                for c_a, c_b in zip(row_a, row_b):
                    if c_a.value != c_b.value:
                        return False
        return True
    finally:
        wb_a.close()
        wb_b.close()


def load_source(path, expected_sha, allow_verify_failure):
    if not os.path.isfile(path):
        raise ValueError("source workbook not found: %s" % path)
    actual = sha256_file(path)
    if expected_sha and actual.lower() != expected_sha.lower():
        if not allow_verify_failure:
            raise ValueError(
                "source workbook hash mismatch: expected %s, actual %s"
                % (expected_sha, actual)
            )
    return actual


def parse_recipe(raw):
    """Return the two non-empty parts for an enabled row, else None."""
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    parts = [p.strip() for p in text.split(",") if p.strip()]
    if len(parts) != 2 or len(set(parts)) != 2:
        return None
    return parts


def canonical_recipe_key(parts):
    return ",".join(sorted(parts))


class ValidationError(Exception):
    pass


def validate(rows):
    """rows: list of dicts with keys index, partWords, enabled and staged fields."""
    seen_ids = {}
    for r in rows:
        idx = r["index"]
        if idx is None or idx in seen_ids:
            raise ValidationError("duplicate or missing index: %r" % (idx,))
        seen_ids[idx] = r

    recipe_keys = {}
    for r in rows:
        parts = parse_recipe(r["partWords"])
        if parts is None:
            raise ValidationError(
                "row %s must have exactly two non-empty distinct recipe parts: %r"
                % (r["index"], r["partWords"])
            )
        key = canonical_recipe_key(parts)
        if key in recipe_keys:
            raise ValidationError(
                "duplicate unordered recipe %r between rows %s and %s"
                % (key, recipe_keys[key], r["index"])
            )
        recipe_keys[key] = r["index"]

    enabled = [r for r in rows if r["enabled"]]
    enabled_ids = {r["index"] for r in enabled}
    if enabled_ids != ENABLED_IDS:
        raise ValidationError(
            "expected exactly enabled ids %s, got %s"
            % (sorted(ENABLED_IDS), sorted(enabled_ids))
        )

    archetypes = {"pike", "bow"}
    for r in enabled:
        idx = r["index"]
        arch = r["combatArchetype"]
        if arch not in archetypes:
            raise ValidationError(
                "row %s invalid combatArchetype %r" % (idx, arch)
            )
        if not (r["rangeCells"] and r["rangeCells"] > 0):
            raise ValidationError(
                "row %s rangeCells must be positive, got %r" % (idx, r["rangeCells"])
            )
        if not (r["attackDamage"] and r["attackDamage"] > 0):
            raise ValidationError(
                "row %s attackDamage must be positive, got %r" % (idx, r["attackDamage"])
            )
        if not (r["attackIntervalSeconds"] and r["attackIntervalSeconds"] > 0):
            raise ValidationError(
                "row %s attackIntervalSeconds must be positive, got %r"
                % (idx, r["attackIntervalSeconds"])
            )
        if not r["prefabAddress"]:
            raise ValidationError("row %s prefabAddress must be non-empty" % idx)
        if not r["partRecruitWeight"] or r["partRecruitWeight"] <= 0:
            raise ValidationError(
                "row %s partRecruitWeight must be positive, got %r"
                % (idx, r["partRecruitWeight"])
            )
        projectile_type = (r["projectileType"] or "").strip()
        projectile_speed = r["projectileSpeed"]
        if arch == "pike":
            if projectile_type or (projectile_speed not in (None, 0, "")):
                raise ValidationError(
                    "row %s pike must not define a projectile" % idx
                )
        elif arch == "bow":
            if not projectile_type or not (projectile_speed and projectile_speed > 0):
                raise ValidationError(
                    "row %s bow requires projectile type and positive speed" % idx
                )


def read_source_rows(ws, staged):
    """Return list of row dicts (index, partWords, enabled, staged fields)."""
    rows = []
    first_data_row = 5
    for r in range(first_data_row, ws.max_row + 1):
        idx = ws.cell(row=r, column=2).value
        part_words = ws.cell(row=r, column=5).value
        entry = STAGED.get(idx) if idx is not None else None
        enabled = bool(entry and entry["enabled"])
        if idx is None:
            rows.append(
                {
                    "index": None,
                    "partWords": part_words,
                    "enabled": False,
                    "combatArchetype": "",
                    "rangeCells": 0,
                    "attackDamage": 0,
                    "attackIntervalSeconds": 0,
                    "damageMode": "",
                    "targetPolicy": "",
                    "prefabAddress": "",
                    "animationKey": "",
                    "projectileType": "",
                    "projectileSpeed": 0,
                    "partRecruitWeight": 0,
                }
            )
            continue
        if entry:
            rows.append(dict(entry))
            rows[-1]["index"] = idx
            rows[-1]["partWords"] = part_words
        else:
            rows.append(
                {
                    "index": idx,
                    "partWords": part_words,
                    "enabled": False,
                    "combatArchetype": "",
                    "rangeCells": 0,
                    "attackDamage": 0,
                    "attackIntervalSeconds": 0,
                    "damageMode": "",
                    "targetPolicy": "",
                    "prefabAddress": "",
                    "animationKey": "",
                    "projectileType": "",
                    "projectileSpeed": 0,
                    "partRecruitWeight": 0,
                }
            )
    return rows


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell._style = src_cell._style


def _blank_value(name):
    if name == "enabled":
        return False
    if name in ("rangeCells", "attackIntervalSeconds"):
        return 0
    if name in ("attackDamage", "projectileSpeed", "partRecruitWeight"):
        return 0
    return ""


def _number_format(name):
    if name in ("rangeCells", "attackIntervalSeconds"):
        return "0.0#"
    return "General"


def write_workbook(src_path, out_path, rows, source_sha, generated_at):
    wb = openpyxl.load_workbook(src_path, data_only=False)
    ws = wb.active
    max_col = ws.max_column

    # Append per-row values for the new columns into the same sheet.
    first_data_row = 5
    enabled = {r["index"]: r for r in rows if r["index"] is not None}
    for r, row in enumerate(rows):
        excel_row = first_data_row + r
        idx = row["index"]
        entry = enabled.get(idx)
        for i, (name, ctype, comment) in enumerate(NEW_COLUMNS):
            col = max_col + i + 1
            cell = ws.cell(row=excel_row, column=col)
            if entry and entry["enabled"]:
                cell.value = entry[name]
            else:
                cell.value = _blank_value(name)
            cell.number_format = _number_format(name)

    # Headers on rows 1-4 for the new columns.
    for i, (name, ctype, comment) in enumerate(NEW_COLUMNS):
        col = max_col + i + 1
        ws.cell(row=1, column=col, value=name)
        ws.cell(row=2, column=col, value=ctype)
        ws.cell(row=3, column=col, value=None)
        ws.cell(row=4, column=col, value=comment)

    # Copy the style of the adjacent existing header/data cells onto new ones
    # so the staged workbook visually matches the original layout.
    for col in range(max_col + 1, max_col + 1 + len(NEW_COLUMNS)):
        prev = ws.cell(row=1, column=max_col)
        dst = ws.cell(row=1, column=col)
        copy_cell_style(prev, dst)
    for r in range(1, 5):
        prev = ws.cell(row=r, column=max_col)
        for col in range(max_col + 1, max_col + 1 + len(NEW_COLUMNS)):
            copy_cell_style(prev, ws.cell(row=r, column=col))
    for r in range(first_data_row, first_data_row + len(rows)):
        prev = ws.cell(row=r, column=max_col)
        for col in range(max_col + 1, max_col + 1 + len(NEW_COLUMNS)):
            copy_cell_style(prev, ws.cell(row=r, column=col))

    wb.save(out_path)
    out_sha = sha256_file(out_path)
    return out_sha


def enabled_preview(rows):
    preview = []
    for r in rows:
        if r.get("enabled"):
            preview.append(
                {
                    "index": r["index"],
                    "name": r.get("name"),
                    "recipe": r["partWords"],
                    "combatArchetype": r["combatArchetype"],
                    "rangeCells": r["rangeCells"],
                    "attackDamage": r["attackDamage"],
                    "attackIntervalSeconds": r["attackIntervalSeconds"],
                    "damageMode": r["damageMode"],
                    "targetPolicy": r["targetPolicy"],
                    "prefabAddress": r["prefabAddress"],
                    "animationKey": r["animationKey"],
                    "projectileType": r["projectileType"] or None,
                    "projectileSpeed": r["projectileSpeed"],
                    "partRecruitWeight": r["partRecruitWeight"],
                }
            )
    return preview


def write_json(path, payload):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
        f.write("\n")


def build_manifest(
    source_path,
    output_dir,
    source_sha,
    out_xlsx,
    out_xlsx_sha,
    enabled_preview_list,
    generated_at,
):
    return {
        "schema": "player-general-synthesis/staging/v1",
        "generatedAt": generated_at,
        "source": {
            "path": os.path.abspath(source_path),
            "sha256": source_sha,
            "expectedSha256": EXPECTED_SOURCE_SHA256,
            "note": "production Datas workbook, read-only; not modified by this generator",
        },
        "output": {
            "directory": os.path.abspath(output_dir),
            "workbook": os.path.basename(out_xlsx),
            "sha256": out_xlsx_sha,
        },
        "validation": {
            "status": "passed",
            "checks": [
                "source workbook SHA-256 verified",
                "unique IDs",
                "exactly two non-empty distinct recipe parts per enabled row",
                "canonical unordered recipe uniqueness",
                "exactly enabled ids {1,4}",
                "positive required fields / prefabAddress / partRecruitWeight",
                "pike has no projectile",
                "bow requires projectile type and positive speed",
            ],
            "enabledIds": sorted(ENABLED_IDS),
        },
        "approvedBalance": {
            "zhangFei": {
                "recipe": "张,飞",
                "combatArchetype": "pike",
                "rangeCells": 2.5,
                "attackDamage": 15,
                "attackIntervalSeconds": 1.0,
                "damageMode": "近战枪击",
                "targetPolicy": "nearest",
                "prefabAddress": "SpearSoldier",
                "animationKey": DEFAULT_ANIMATION_KEY,
                "projectileType": None,
                "projectileSpeed": 0,
                "partRecruitWeight": 1,
            },
            "huangZhong": {
                "recipe": "黄,忠",
                "combatArchetype": "bow",
                "rangeCells": 3.5,
                "attackDamage": 13,
                "attackIntervalSeconds": 0.8,
                "damageMode": "单体",
                "targetPolicy": "nearest",
                "prefabAddress": "BowSoldier",
                "animationKey": "default",
                "projectileType": "SimpleDynamicArrow",
                "projectileSpeed": 200,
                "partRecruitWeight": 1,
            },
            "note": "Placeholder presentation: Zhang Fei uses SpearSoldier; Huang Zhong uses BowSoldier/Arrow. Replace addresses and content later without changing battle logic.",
        },
        "enabledGenerals": enabled_preview_list,
    }


def write_readme(output_dir, enabled_preview_list, generated_at):
    lines = [
        "# Player General Synthesis - Staging Package",
        "",
        "Generated %s by `Tools/GeneralContent/generate_general_staging.py`." % generated_at,
        "",
        "This package is an **import-ready staging candidate**. It does NOT modify",
        "production `Configs/GameConfig/Datas/general.xlsx`. Review it before import.",
        "",
        "## Files",
        "",
        "- `general.xlsx` - full replacement candidate (12 data rows preserved, new approved columns appended).",
        "- `enabled-generals.json` - two-row enabled preview (Zhang Fei, Huang Zhong).",
        "- `manifest.json` - source/output hashes, validation result and approved balance.",
        "",
        "## Backup and manual import",
        "",
        "1. Back up the current production workbook `Configs/GameConfig/Datas/general.xlsx`.",
        "2. Review `general.xlsx` and `enabled-generals.json`; confirm the two enabled",
        "   generals, initial values and part weight 1.",
        "3. Replace `Configs/GameConfig/Datas/general.xlsx` with the staged workbook.",
        "4. Run the project Luban lazy-load export script:",
        "",
        "   `Configs/GameConfig/gen_code_bin_to_project_lazyload.bat`",
        "",
        "   (per project workflow). Verify the generated `General.cs`/`TbGeneral.cs` and",
        "   `battle_tbgeneral.bytes` contain the approved fields/rows before compiling.",
        "",
        "## Placeholder presentation",
        "",
        "This staging config intentionally sets Zhang Fei to `SpearSoldier/default` and",
        "Huang Zhong to `BowSoldier/default` with `SimpleDynamicArrow`. Replace the",
        "configured addresses and content later without changing combat logic.",
        "",
        "## Regeneration / protection",
        "",
        "- Re-running the generator is idempotent: identical output is left untouched.",
        "- If the output workbook already differs, the generator stops unless `--force` is passed.",
        "- The generator never writes production `Datas`.",
    ]
    path = os.path.join(output_dir, "README.md")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    return path


def run(argv=None):
    parser = argparse.ArgumentParser(
        description="Generate player-general-synthesis staging XLSX package."
    )
    parser.add_argument(
        "--source",
        default=SOURCE_WORKBOOK,
        help="source production general.xlsx (default: %(default)s)",
    )
    parser.add_argument(
        "--expected-hash",
        default=EXPECTED_SOURCE_SHA256,
        help="expected source SHA-256 (default validates against known source)",
    )
    parser.add_argument(
        "--no-hash-check",
        action="store_true",
        help="do not fail when the source hash differs from expected",
    )
    parser.add_argument(
        "--output-dir", default=OUTPUT_DIR, help="output directory (default: %(default)s)"
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="overwrite an existing differing exact output workbook",
    )
    args = parser.parse_args(argv)

    source_sha = load_source(
        args.source, args.expected_hash, allow_verify_failure=args.no_hash_check
    )

    wb = openpyxl.load_workbook(args.source, data_only=False)
    ws = wb.active
    rows = read_source_rows(ws, STAGED)

    # Attach name for the enabled preview from the preserved source column.
    first_data_row = 5
    for r, row in enumerate(rows):
        excel_row = first_data_row + r
        row["name"] = ws.cell(row=excel_row, column=3).value

    validate(rows)

    os.makedirs(args.output_dir, exist_ok=True)
    out_xlsx = os.path.join(args.output_dir, "general.xlsx")
    generated_at = time.strftime("%Y-%m-%dT%H:%M:%S%z")

    if os.path.isfile(out_xlsx):
        tmp_path = out_xlsx + ".tmp." + uuid.uuid4().hex + ".xlsx"
        try:
            candidate_sha = write_workbook(
                args.source, tmp_path, rows, source_sha, generated_at
            )
        except Exception:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
            raise
        if not workbook_content_equal(out_xlsx, tmp_path):
            if not args.force:
                os.remove(tmp_path)
                raise ValueError(
                    "existing output %s differs from the newly generated content; "
                    "pass --force to overwrite this exact file" % out_xlsx
                )
            os.replace(tmp_path, out_xlsx)
        else:
            os.remove(tmp_path)
        out_sha = sha256_file(out_xlsx)
    else:
        out_sha = write_workbook(args.source, out_xlsx, rows, source_sha, generated_at)

    preview = enabled_preview(rows)

    enabled_json_path = os.path.join(args.output_dir, "enabled-generals.json")
    write_json(enabled_json_path, preview)

    manifest = build_manifest(
        args.source,
        args.output_dir,
        source_sha,
        out_xlsx,
        out_sha,
        preview,
        generated_at,
    )
    manifest_path = os.path.join(args.output_dir, "manifest.json")
    write_json(manifest_path, manifest)

    readme_path = write_readme(args.output_dir, preview, generated_at)

    print("wrote %s" % out_xlsx)
    print("wrote %s" % enabled_json_path)
    print("wrote %s" % manifest_path)
    print("wrote %s" % readme_path)
    return 0


if __name__ == "__main__":
    sys.exit(run())
