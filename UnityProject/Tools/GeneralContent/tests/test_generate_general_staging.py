#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Focused unittest for the player-general-synthesis staging generator."""

import hashlib
import json
import os
import shutil
import sys
import tempfile
import unittest

import openpyxl

sys.path.insert(
    0, os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
)
import generate_general_staging as gen


def sha256_bytes(data):
    return hashlib.sha256(data).hexdigest()


def make_source_workbook(path, rows, new_cols_values=None):
    """Build a Luban-style general.xlsx with the same shape as production."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["##var", "index", "name", "family", "partWords", "weaponType", "status"]
    types = ["##type", "int", "string", "string", "(list#sep=,),string", "int", "string?"]
    ws.append(headers)
    ws.append(types)
    ws.append(["##group"] + [None] * 6)
    ws.append(["##"] + ["索引(主键)", "名称", "姓氏", "名字拆字", "武器类型", "状态"])
    for r in rows:
        ws.append([None] + list(r))
    wb.save(path)
    return path


class GeneratorTest(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.addCleanup(shutil.rmtree, self.tmp)
        self.src = os.path.join(self.tmp, "general.xlsx")
        # Same 12 rows as production (index/name/family/partWords/weaponType/status).
        rows = [
            (0, "赵云", "赵", "赵,云", 1, "PARTIAL_CORE_CONFIG"),
            (1, "张飞", "张", "张,飞", 1, "PARTIAL_CORE_CONFIG"),
            (2, "马超", "马", "马,超", 1, "PARTIAL_CORE_CONFIG"),
            (3, "关羽", "关", "关,羽", 2, "PARTIAL_CORE_CONFIG"),
            (4, "黄忠", "黄", "黄,忠", 0, "PARTIAL_CORE_CONFIG"),
            (5, "关平", "关", "关,平", 2, "PARTIAL_CORE_CONFIG"),
            (6, "关兴", "关", "关,兴", 2, "PARTIAL_CORE_CONFIG"),
            (7, "张苞", "张", "张,苞", 1, "PARTIAL_CORE_CONFIG"),
            (8, "张翼", "张", "张,翼", 3, "PARTIAL_CORE_CONFIG"),
            (9, "黄盖", "黄", "黄,盖", 3, "PARTIAL_CORE_CONFIG"),
            (10, "刘备", "刘", "刘,备", 3, "PARTIAL_CORE_CONFIG"),
            (11, "黄祖", "黄", "黄,祖", 0, "PARTIAL_CORE_CONFIG"),
        ]
        make_source_workbook(self.src, rows)

    def source_sha(self):
        with open(self.src, "rb") as f:
            return sha256_bytes(f.read())

    def load_rows(self, ws):
        rows = []
        for r in range(5, ws.max_row + 1):
            idx = ws.cell(row=r, column=2).value
            if idx is None:
                continue
            rows.append(
                {
                    "index": idx,
                    "partWords": ws.cell(row=r, column=5).value,
                    "enabled": ws.cell(row=r, column=8).value,
                }
            )
        return rows

    def test_validate_accepts_valid_staging(self):
        rows = gen.read_source_rows(openpyxl.load_workbook(self.src).active, gen.STAGED)
        for r in rows:
            r["name"] = "n"
        gen.validate(rows)
        enabled = [r for r in rows if r["enabled"]]
        self.assertEqual({r["index"] for r in enabled}, {1, 4})

    def test_validate_rejects_extra_enabled(self):
        rows = gen.read_source_rows(openpyxl.load_workbook(self.src).active, gen.STAGED)
        for r in rows:
            r["name"] = "n"
        rows[0]["enabled"] = True
        rows[0].update(
            {
                "combatArchetype": "pike",
                "rangeCells": 1,
                "attackDamage": 1,
                "attackIntervalSeconds": 1,
                "prefabAddress": "x",
                "partRecruitWeight": 1,
                "projectileType": "",
                "projectileSpeed": 0,
            }
        )
        with self.assertRaises(gen.ValidationError):
            gen.validate(rows)

    def test_validate_rejects_duplicate_recipe(self):
        rows = gen.read_source_rows(openpyxl.load_workbook(self.src).active, gen.STAGED)
        for r in rows:
            r["name"] = "n"
        rows[0]["partWords"] = "张,飞"  # duplicate canonical key with row 1
        with self.assertRaises(gen.ValidationError):
            gen.validate(rows)

    def test_validate_rejects_malformed_recipe(self):
        rows = gen.read_source_rows(openpyxl.load_workbook(self.src).active, gen.STAGED)
        for r in rows:
            r["name"] = "n"
        rows[0]["partWords"] = "张,张"  # not two distinct parts
        with self.assertRaises(gen.ValidationError):
            gen.validate(rows)

    def test_validate_pike_forbids_projectile(self):
        rows = gen.read_source_rows(openpyxl.load_workbook(self.src).active, gen.STAGED)
        rows[0]["enabled"] = True
        rows[0]["combatArchetype"] = "pike"
        rows[0]["rangeCells"] = 1
        rows[0]["attackDamage"] = 1
        rows[0]["attackIntervalSeconds"] = 1
        rows[0]["prefabAddress"] = "x"
        rows[0]["partRecruitWeight"] = 1
        rows[0]["projectileType"] = "SimpleDynamicArrow"
        rows[0]["projectileSpeed"] = 200
        with self.assertRaises(gen.ValidationError):
            gen.validate(rows)

    def test_validate_bow_requires_projectile(self):
        rows = gen.read_source_rows(openpyxl.load_workbook(self.src).active, gen.STAGED)
        rows[0]["enabled"] = True
        rows[0]["combatArchetype"] = "bow"
        rows[0]["rangeCells"] = 1
        rows[0]["attackDamage"] = 1
        rows[0]["attackIntervalSeconds"] = 1
        rows[0]["prefabAddress"] = "x"
        rows[0]["partRecruitWeight"] = 1
        rows[0]["projectileType"] = ""
        rows[0]["projectileSpeed"] = 0
        with self.assertRaises(gen.ValidationError):
            gen.validate(rows)

    def test_source_hash_mismatch_rejected_by_default(self):
        with self.assertRaises(ValueError):
            gen.load_source(self.src, "0" * 64, allow_verify_failure=False)

    def test_source_hash_ok(self):
        actual = self.source_sha()
        self.assertEqual(gen.load_source(self.src, actual, False), actual)

    def test_end_to_end_generation_preserves_rows_and_enabled(self):
        out = os.path.join(self.tmp, "out")
        os.makedirs(out)
        source_sha = self.source_sha()
        rc = gen.run(
            [
                "--source", self.src,
                "--expected-hash", source_sha,
                "--output-dir", out,
            ]
        )
        self.assertEqual(rc, 0)
        wb = openpyxl.load_workbook(os.path.join(out, "general.xlsx"))
        ws = wb.active
        rows = self.load_rows(ws)
        self.assertEqual(len(rows), 12)
        enabled = [r for r in rows if r["enabled"]]
        self.assertEqual({r["index"] for r in enabled}, {1, 4})
        self.assertEqual(
            ws.cell(row=4, column=15).value,
            "当前枪兵/弓兵回退地址，后续可替换为专属 Prefab",
        )

        # Zhang Fei row values.
        zf = next(r for r in rows if r["index"] == 1)
        self.assertEqual(ws.cell(row=6, column=9).value, "pike")
        self.assertEqual(ws.cell(row=6, column=10).value, 2.5)
        self.assertEqual(ws.cell(row=6, column=11).value, 15)
        self.assertEqual(ws.cell(row=6, column=12).value, 1.0)
        self.assertEqual(ws.cell(row=6, column=13).value, "近战枪击")
        self.assertEqual(ws.cell(row=6, column=14).value, "nearest")
        self.assertEqual(ws.cell(row=6, column=15).value, "SpearSoldier")
        self.assertEqual(ws.cell(row=6, column=16).value, "default")
        self.assertIn(ws.cell(row=6, column=17).value, (None, ""))
        self.assertEqual(ws.cell(row=6, column=18).value, 0)
        self.assertEqual(ws.cell(row=6, column=19).value, 1)

        # Huang Zhong row values.
        hz_row = 9
        self.assertEqual(ws.cell(row=hz_row, column=9).value, "bow")
        self.assertEqual(ws.cell(row=hz_row, column=10).value, 3.5)
        self.assertEqual(ws.cell(row=hz_row, column=11).value, 13)
        self.assertEqual(ws.cell(row=hz_row, column=12).value, 0.8)
        self.assertEqual(ws.cell(row=hz_row, column=13).value, "单体")
        self.assertEqual(ws.cell(row=hz_row, column=15).value, "BowSoldier")
        self.assertEqual(ws.cell(row=hz_row, column=16).value, "default")
        self.assertEqual(ws.cell(row=hz_row, column=17).value, "SimpleDynamicArrow")
        self.assertEqual(ws.cell(row=hz_row, column=18).value, 200)
        self.assertEqual(ws.cell(row=hz_row, column=19).value, 1)

        # Disabled rows keep original data, enabled=False, blank new fields.
        for idx in (0, 2, 3, 5, 6, 7, 8, 9, 10, 11):
            r = next(x for x in rows if x["index"] == idx)
            self.assertFalse(r["enabled"])

        # Artifacts present.
        self.assertTrue(os.path.isfile(os.path.join(out, "enabled-generals.json")))
        self.assertTrue(os.path.isfile(os.path.join(out, "manifest.json")))
        self.assertTrue(os.path.isfile(os.path.join(out, "README.md")))
        with open(os.path.join(out, "enabled-generals.json"), encoding="utf-8") as f:
            preview = json.load(f)
        self.assertEqual({p["index"] for p in preview}, {1, 4})
        with open(os.path.join(out, "manifest.json"), encoding="utf-8") as f:
            manifest = json.load(f)
        self.assertEqual(manifest["validation"]["status"], "passed")
        self.assertEqual(manifest["source"]["sha256"], source_sha)
        self.assertEqual(manifest["source"]["path"], os.path.abspath(self.src))
        self.assertEqual(manifest["output"]["directory"], os.path.abspath(out))
        self.assertEqual(
            manifest["output"]["sha256"],
            gen.sha256_file(os.path.join(out, "general.xlsx")),
        )

    def test_rerun_is_idempotent_and_differing_output_protected(self):
        out = os.path.join(self.tmp, "out")
        os.makedirs(out)
        source_sha = self.source_sha()
        gen.run(["--source", self.src, "--expected-hash", source_sha, "--output-dir", out])
        xlsx = os.path.join(out, "general.xlsx")
        with open(xlsx, "rb") as f:
            before = f.read()

        # Identical rerun succeeds and leaves the exact file unchanged (no rewrite).
        gen.run(["--source", self.src, "--expected-hash", source_sha, "--output-dir", out])
        with open(xlsx, "rb") as f:
            after = f.read()
        self.assertEqual(before, after)

        # Capture a fresh reference copy for later content comparison.
        ref = os.path.join(out, "ref.xlsx")
        shutil.copyfile(xlsx, ref)

        # Differing output is protected without --force.
        wb = openpyxl.load_workbook(xlsx)
        wb.active.cell(row=6, column=9).value = "CHANGED"
        wb.save(xlsx)
        wb.close()
        with self.assertRaises(ValueError):
            gen.run(["--source", self.src, "--expected-hash", source_sha, "--output-dir", out])

        # --force overwrites the differing exact file (content restored).
        gen.run(
            ["--source", self.src, "--expected-hash", source_sha, "--output-dir", out, "--force"]
        )
        self.assertTrue(gen.workbook_content_equal(xlsx, ref))

    def test_never_writes_production(self):
        # Confirms the generator only ever targets OUTPUT_DIR under outputs/.
        self.assertTrue("Datas" not in gen.OUTPUT_DIR)
        self.assertTrue(gen.OUTPUT_DIR.startswith(gen.PROJECT_ROOT))


if __name__ == "__main__":
    unittest.main(verbosity=2)
